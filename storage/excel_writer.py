import os
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from config import EXCEL_FILE


HEADERS = ["Источник", "Название", "Цена", "Город", "Ссылка", "Дата парсинга"]
HEADER_COLOR = "1A1A2E"
ROW_COLORS = ["FFFFFF", "F5F5F5"]
SOURCE_COLORS = {
    "OLX.kg": "E8F5E9",
    "Lalafo.kg": "E3F2FD",
}


def _style_header(ws):
    fill = PatternFill("solid", fgColor=HEADER_COLOR)
    font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        bottom=Side(style="medium", color="000000")
    )
    for col, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border


def _get_row_fill(source: str, row_idx: int) -> PatternFill:
    for key, color in SOURCE_COLORS.items():
        if key in source:
            return PatternFill("solid", fgColor=color)
    color = ROW_COLORS[row_idx % 2]
    return PatternFill("solid", fgColor=color)


def save_to_excel(all_items: list[dict]) -> str:
    """Сохраняет данные в Excel. Если файл существует — добавляет новый лист."""
    if os.path.exists(EXCEL_FILE):
        wb = openpyxl.load_workbook(EXCEL_FILE)
    else:
        wb = openpyxl.Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    sheet_name = datetime.now().strftime("%d.%m.%Y")
    # Если лист с таким именем уже есть — удаляем старый
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

    ws = wb.create_sheet(title=sheet_name)

    _style_header(ws)

    for i, item in enumerate(all_items, 2):
        row_data = [
            item.get("source", ""),
            item.get("title", ""),
            item.get("price", ""),
            item.get("location", ""),
            item.get("link", ""),
            item.get("parsed_at", ""),
        ]
        fill = _get_row_fill(item.get("source", ""), i)
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=i, column=col, value=value)
            cell.fill = fill
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            if col == 5:  # ссылка
                cell.font = Font(color="0066CC", underline="single")

    # Авто-ширина колонок
    col_widths = [15, 50, 15, 20, 50, 18]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.row_dimensions[1].height = 25
    ws.freeze_panes = "A2"

    wb.save(EXCEL_FILE)
    print(f"[Excel] Сохранено {len(all_items)} строк → {EXCEL_FILE}")
    return EXCEL_FILE
